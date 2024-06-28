from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from time import sleep

paginaFecha=[]
lePlan=openpyxl.load_workbook('dados_clientes.xlsx')
paginaClientes=lePlan['Sheet1']
driver=webdriver.Firefox()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in paginaClientes.iter_rows(min_row=2,values_only=True):
    nome, valor, cpf, vencimento = linha
    
    
    #input cpfInput
    sleep(3)
    campoPesquisa=driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    campoPesquisa.clear()
    sleep(1)
    campoPesquisa.send_keys(cpf)
    sleep(1)
    btnPesquisar=driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    btnPesquisar.click()
    sleep(4)
    statusLabel=driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if statusLabel.text== 'em dia':
        fechaPlani=openpyxl.load_workbook('planilha_fechamento.xlsx')
        paginaFecha=fechaPlani['Sheet1']
        dataPag=driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodoPag=driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
        sodata=dataPag.text.split()[3]
        sometodo=metodoPag.text.split()[3]
        
        paginaFecha.append([nome,valor,cpf,vencimento,'em dia',sodata,sometodo])
        fechaPlani.save('planilha_fechamento.xlsx')
    else:
        fechaPlani=openpyxl.load_workbook('planilha_fechamento.xlsx')
        paginaFecha=fechaPlani['Sheet1']
        
        paginaFecha.append([nome,valor,cpf,vencimento,'pendente'])
        fechaPlani.save('planilha_fechamento.xlsx')
    
    